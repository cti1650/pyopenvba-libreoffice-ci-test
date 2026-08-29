#!/usr/bin/env python3
"""
Create an Excel file with VBA macros using pyOpenVBA.
This script demonstrates how to programmatically add VBA code to Excel files.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from pyopenvba import ExcelFile


def create_base_excel(output_path: Path) -> None:
    """Create a base Excel file with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Test Data"
    ws["A2"] = 10
    ws["B2"] = 20
    # Save as .xlsx first
    xlsx_path = output_path.with_suffix(".xlsx")
    wb.save(xlsx_path)
    print(f"Created base Excel file: {xlsx_path}")
    return xlsx_path


def inject_vba(xlsx_path: Path, vba_code: str, output_path: Path) -> None:
    """Inject VBA code into Excel file and save as .xlsm."""
    # pyOpenVBA can create a new xlsm file
    with ExcelFile.create_new(str(output_path)) as wb:
        # Get existing modules (new files have default modules)
        modules = wb.module_names()
        print(f"Available modules: {modules}")

        # Update Module1 if it exists, otherwise try creating new
        try:
            if "Module1" in modules:
                wb.set_module("Module1", vba_code)
                print("Updated Module1 with VBA code")
            else:
                wb.set_module("SampleModule", vba_code)
                print("Created SampleModule with VBA code")
        except KeyError as e:
            print(f"Warning: Could not set module: {e}")
            print("Saving file with default modules only")

        wb.save()
    print(f"Created Excel file with VBA: {output_path}")


def read_vba_file(vba_path: Path) -> str:
    """Read VBA file and remove Attribute line if present."""
    vba_code = vba_path.read_text(encoding="utf-8")
    lines = vba_code.split("\n")
    if lines[0].startswith("Attribute VB_Name"):
        vba_code = "\n".join(lines[1:])
    return vba_code


def create_win32api_excel(output_path: Path, vba_code: str) -> None:
    """Create Excel file with Win32 API VBA module."""
    with ExcelFile.create_new(str(output_path)) as wb:
        modules = wb.module_names()
        print(f"Available modules for Win32 API file: {modules}")

        try:
            if "Module1" in modules:
                wb.set_module("Module1", vba_code)
                print("Updated Module1 with Win32 API code")
            else:
                wb.set_module("Win32ApiModule", vba_code)
                print("Created Win32ApiModule")
        except KeyError as e:
            print(f"Warning: Could not set Win32 API module: {e}")

        wb.save()
    print(f"Created Win32 API Excel file: {output_path}")


def verify_vba_in_file(xlsm_path: Path) -> None:
    """Verify VBA code was injected and display summary."""
    print(f"\n--- Verifying: {xlsm_path.name} ---")

    if not xlsm_path.exists():
        print(f"  ERROR: File not found")
        return

    print(f"  File size: {xlsm_path.stat().st_size} bytes")

    try:
        with ExcelFile(str(xlsm_path)) as wb:
            modules = wb.module_names()
            print(f"  Modules: {modules}")

            for module_name in modules:
                try:
                    source = wb.get_module(module_name)
                    lines = source.strip().split("\n")
                    print(f"\n  [{module_name}] ({len(lines)} lines)")

                    # Check for Win32 API declarations
                    declare_count = sum(1 for line in lines if "Declare" in line)
                    function_count = sum(1 for line in lines if line.strip().startswith(("Public Function", "Public Sub", "Private Function", "Private Sub")))

                    if declare_count > 0:
                        print(f"    Win32 API Declarations: {declare_count}")
                    if function_count > 0:
                        print(f"    Functions/Subs: {function_count}")

                    # Show first few lines
                    print("    Preview:")
                    for line in lines[:5]:
                        if line.strip():
                            print(f"      {line[:60]}{'...' if len(line) > 60 else ''}")
                    if len(lines) > 5:
                        print(f"      ... ({len(lines) - 5} more lines)")

                except Exception as e:
                    print(f"    Could not read: {e}")

    except Exception as e:
        print(f"  ERROR reading file: {e}")


def main():
    # Paths
    project_root = Path(__file__).parent.parent
    output_dir = project_root / "output"
    output_dir.mkdir(exist_ok=True)

    # Create basic VBA Excel file
    print("=" * 50)
    print("Creating basic VBA Excel file")
    print("=" * 50)
    vba_path = project_root / "vba" / "sample_module.bas"
    output_xlsm = output_dir / "test_workbook.xlsm"

    vba_code = read_vba_file(vba_path)
    inject_vba(None, vba_code, output_xlsm)

    # Create Win32 API VBA Excel file
    print("\n" + "=" * 50)
    print("Creating Win32 API VBA Excel file")
    print("=" * 50)
    win32_vba_path = project_root / "vba" / "win32api_module.bas"
    win32_output_xlsm = output_dir / "win32api_workbook.xlsm"

    if win32_vba_path.exists():
        win32_vba_code = read_vba_file(win32_vba_path)
        create_win32api_excel(win32_output_xlsm, win32_vba_code)
    else:
        print(f"Win32 API VBA file not found: {win32_vba_path}")

    # Verify created files
    print("\n" + "=" * 50)
    print("Verification Results")
    print("=" * 50)
    verify_vba_in_file(output_xlsm)
    verify_vba_in_file(win32_output_xlsm)

    # Summary
    print("\n" + "=" * 50)
    print("Summary")
    print("=" * 50)
    for f in output_dir.iterdir():
        if f.suffix == ".xlsm":
            print(f"  {f.name}: {f.stat().st_size} bytes")

    print("\nDone!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
